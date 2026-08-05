from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from typing import Optional
import re
import unicodedata

import pdfplumber
import openpyxl
from bs4 import BeautifulSoup

MONTH_ES = {
    'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4, 'may': 5, 'jun': 6,
    'jul': 7, 'ago': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dic': 12,
}

CATEGORIES = [
    'Alimentación',
    'Restaurantes',
    'Ropa/Compras',
    'Taxi',
    'Coche',
    'Transporte',
    'Cultura/Entretenimiento',
    'Suscripciones/Tech',
    'Hogar/Recibos',
    'Salud',
    'Efectivo',
    'Devolución',
    'Bizum',
    'Apuestas',
    'Clubs',
    'Formación',
    'Regalos',
    'Viajes',
    'Impuestos',
    'Boda',
    'Hipoteca',
    'Compra vivienda',
    'Otros',
]

CATEGORY_EMOJI = {
    'Alimentación': '🛒',
    'Restaurantes': '🍽',
    'Ropa/Compras': '👔',
    'Taxi': '🚕',
    'Coche': '🚗',
    'Transporte': '🚌',
    'Cultura/Entretenimiento': '🎭',
    'Suscripciones/Tech': '💻',
    'Hogar/Recibos': '🏠',
    'Salud': '❤',
    'Efectivo': '💵',
    'Bizum': '📲',
    'Apuestas': '🎲',
    'Clubs': '🏛',
    'Formación': '📚',
    'Regalos': '🎁',
    'Viajes': '✈️',
    'Impuestos': '🏛',
    'Devolución': '↩️',
    'Hipoteca': '🏦',
    'Compra vivienda': '🏡',
    'Otros': '📦',
}

# Keywords that identify own accounts (to filter internal transfers)
OWN_ACCOUNT_KEYWORDS = [
    'CAVALLER GRAU PABLO',
    'ES2000730100520612209683',  # Openbank IBAN
    'ES3615860001470793034611',  # Trade Republic IBAN
    'TRADE ES',
    'IBKR',
    'PABLO CAVALLER GRAU',
    'MYINVESTOR',
    'BINANCE',
    'BIFINITY',
    'NAGA MARKETS',
    'BGET',
    'BUTGET',
    'TRADE REPUBLIC',
    'PAYOUT TO TRANSIT',
    'PAYOUT TO TRANSIT ACCOUNT',
    'ES3900812709530005501262',  # Sabadell personal account
    'PABLO SABADELL',
    'MAR A ROSALIA',             # María's Revolut/own account alias
    'RUISANCHEZ GONZALEZ',       # María's specific surname combo (internal transfers)
    'SANTANDER MERI',            # María's own Santander account (alias "Meri")
    'BBVA MERI',                 # María's own BBVA-side alias (same as en parsers BBVA)
]


@dataclass
class Transaction:
    date: datetime
    description: str
    amount: float       # always positive; direction determined by tx_type
    tx_type: str        # 'expense' | 'income' | 'internal' | 'investment' | 'patrimonio' | 'cash_withdrawal'
    bank: str
    category: str = ''

    def fmt_date(self) -> str:
        return self.date.strftime('%d/%m/%Y')

    def fmt_amount(self) -> str:
        s = f"{self.amount:,.2f}€"
        return s.replace(',', 'X').replace('.', ',').replace('X', '.')

    @property
    def category_label(self) -> str:
        emoji = CATEGORY_EMOJI.get(self.category, '📦')
        return f"{emoji} {self.category}" if self.category else "❓ Sin categoría"


# ── Helpers ────────────────────────────────────────────────────────────────────

def _parse_amount(s: str) -> Optional[float]:
    if not s:
        return None
    s = re.sub(r'[€\s]', '', str(s))
    s = s.replace('.', '').replace(',', '.')
    try:
        return abs(float(s))
    except ValueError:
        return None


def _parse_date_es(s: str) -> Optional[datetime]:
    s = ' '.join(str(s).split())
    m = re.search(r'(\d{1,2})\s+(\w{3})\s+(\d{4})', s)
    if not m:
        return None
    day, mon, year = m.groups()
    month = MONTH_ES.get(mon.lower())
    if not month:
        return None
    try:
        return datetime(int(year), month, int(day))
    except ValueError:
        return None


def _is_internal(description: str) -> bool:
    d = description.upper()
    return any(k.upper() in d for k in OWN_ACCOUNT_KEYWORDS)


def _strip_accents(s: str) -> str:
    return ''.join(c for c in unicodedata.normalize('NFKD', str(s or '')) if not unicodedata.combining(c))


def disambiguate_duplicates(transactions: list[Transaction]) -> None:
    """Mutates descriptions in-place when two or more transactions in the same
    import batch would collide on the Sheets dedupe key (fecha + descripción +
    importe con signo + banco — ver `SheetsClient.write_transactions` en
    sheets.py) despite being genuinely distinct movements, p.ej. dos cheques
    bancarios idénticos emitidos el mismo día. Sin esto, `write_transactions`
    descartaría el segundo como si fuera un duplicado real."""
    groups: dict[tuple, list[Transaction]] = {}
    for tx in transactions:
        importe = -tx.amount if tx.tx_type == 'expense' else tx.amount
        key = (tx.fmt_date(), tx.description, round(importe, 2), tx.bank)
        groups.setdefault(key, []).append(tx)
    for group in groups.values():
        if len(group) > 1:
            for i, tx in enumerate(group, start=1):
                tx.description = f"{tx.description} ({i}/{len(group)})"


def _clean_tr_desc(s: str) -> str:
    return re.sub(r'\s*null\s*$', '', str(s or ''), flags=re.IGNORECASE).strip()


def _clean_openbank_desc(concepto: str) -> str:
    m = re.match(r'COMPRA EN (.+?)(?:,\s*CON LA TARJETA|$)', concepto, re.I)
    if m:
        return m.group(1).strip()
    m = re.match(r'BIZUM A FAVOR DE (.+?)(?:\s+CONCEPTO|$)', concepto, re.I)
    if m:
        return f"Bizum → {m.group(1).strip()}"
    return concepto


def _classify_openbank(date: datetime, concepto: str, amount: float) -> Optional[Transaction]:
    """Reglas de clasificación compartidas por las tres variantes de parser de
    Openbank (HTML/XLS, PDF 'Extracto', PDF 'Cuentas - Movimientos'). Cada
    variante solo se encarga de extraer (date, concepto, amount) de su propio
    formato de origen; toda la lógica de negocio vive aquí."""
    cu = concepto.upper()

    # Nota: la exclusión de "Psicoterapia Ayuso" (nómina de María) vivía aquí
    # hardcodeada; ahora es una regla 'exclude' en la hoja Reglas, aplicada
    # post-parseo por classifier.apply_exclusions() — ver migrate_rules_to_sheet.py.

    # Salary / reimbursements from DiverInvest
    if 'DIVERINVEST' in cu:
        if amount > 0:
            desc = 'Nómina DiverInvest' if 'NOMINA' in cu or 'NÓMINA' in cu else f'DiverInvest: {concepto}'
            return Transaction(date=date, description=desc, amount=amount,
                               tx_type='income', bank='Openbank')

    # ATM cash withdrawal
    if 'CAJERO' in cu or ('DISPOSICION' in cu and 'CAJERO' in cu):
        return Transaction(date=date, description=f'Cajero {abs(amount):.0f}€',
                           amount=abs(amount), tx_type='cash_withdrawal', bank='Openbank')

    # Revolut top-ups (loading Revolut via Openbank card)
    if 'REVOLUT' in cu and amount < 0:
        return Transaction(date=date, description=concepto, amount=abs(amount),
                           tx_type='internal', bank='Openbank')

    # Internal transfers (own accounts, incl. Trade Republic aliases)
    if (_is_internal(concepto) or 'TRADE REPÚBLIC' in cu
            or 'TRADE REPÚBLICA' in cu or 'TRADE REPUBLIC' in cu):
        return Transaction(date=date, description=concepto, amount=abs(amount),
                           tx_type='internal', bank='Openbank')

    # Card refunds
    if 'ABONO EN LA TARJETA' in cu:
        return Transaction(date=date, description='Devolución tarjeta', amount=abs(amount),
                           tx_type='income', bank='Openbank')

    # Direct debits
    if 'RECIBO' in cu:
        desc = re.sub(r'\s*Nº RECIBO.*', '', concepto, flags=re.I).strip()
        return Transaction(date=date, description=desc, amount=abs(amount),
                           tx_type='expense', bank='Openbank')

    # Expenses (negative amount)
    if amount < 0:
        desc = _clean_openbank_desc(concepto)
        return Transaction(date=date, description=desc, amount=abs(amount),
                           tx_type='expense', bank='Openbank')

    # Income (positive, not already handled)
    if amount > 0:
        if 'BIZUM DE' in cu:
            desc = re.sub(r'^BIZUM DE\s*', '', concepto, flags=re.I)
            desc = re.sub(r'\s+CONCEPTO.*', '', desc, flags=re.I).strip()
            return Transaction(date=date, description=f'Bizum de {desc}',
                               amount=amount, tx_type='income', bank='Openbank')
        return Transaction(date=date, description=concepto, amount=amount,
                           tx_type='income', bank='Openbank')

    return None


def detect_bank(filename: str, content_hint: str = '') -> str:
    import os
    name = os.path.basename(filename).lower()
    name_noaccent = _strip_accents(name)
    # BBVA app exports are named "<algo> - Últimos movimientos.xlsx" con timestamp,
    # sin ninguna palabra reconocible como "bbva" en el nombre — hay que mirar este
    # patrón de nombre antes que el genérico 'movimientos' de Openbank más abajo.
    if name.endswith('.xlsx') and ('ultimos movimientos' in name_noaccent or 'informe bbva' in name_noaccent):
        return 'bbva'
    if 'extracto' in name or 'trade' in name or 'republic' in name:
        return 'trade_republic'
    if 'movimientos' in name or 'openbank' in name or 'cuenta' in name:
        return 'openbank_pdf' if name.endswith('.pdf') else 'openbank'
    if 'revolut' in name or 'account-statement' in name:
        return 'revolut'
    # Content-based detection for files with non-descriptive names
    if content_hint:
        ch = content_hint.upper()
        if '0049' in ch or 'SANTANDER' in ch:
            return 'santander'
        if 'TRADE REPUBLIC' in ch or 'TRBKESM' in ch:
            return 'trade_republic'
        if 'CUENTAS - MOVIMIENTOS' in ch or 'CTA NOMINA OPEN' in ch:
            return 'openbank_cuentas'
        if 'OPENBANK' in ch:
            return 'openbank_pdf'
    return 'unknown'


def _detect_bank_from_pdf(pdf_path: str) -> str:
    """Read first page of PDF and detect bank from content."""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            # Use crop to strip sidebar before text extraction
            page = pdf.pages[0]
            cropped = page.crop((50, 0, page.width, page.height))
            text = cropped.extract_text() or ''
        return detect_bank(pdf_path, content_hint=text)
    except Exception:
        return 'unknown'


# ── Trade Republic PDF Parser ──────────────────────────────────────────────────

class TradeRepublicParser:
    """
    Parse Trade Republic PDFs using word coordinates.
    FECHA/TIPO (x < 149) is split by content (day/month/year vs text); DESC vs
    money columns (x >= 149) is split by whether the token looks like an amount.
    The ENTRADA/SALIDA/BALANCE x-offsets themselves drift between PDF exports, so
    money tokens are ordered by position instead: rightmost = balance, next = amount.
    """

    # x0 column boundaries (from actual PDF word positions)
    X_TIPO = 100
    X_DESC = 135  # tipo continuation words (e.g. "tarjeta") max out ~117; desc starts ~148

    def parse(self, pdf_path: str) -> list[Transaction]:
        in_section = False
        tx_blocks: list[list[dict]] = []
        current_block: list[dict] = []

        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                words = page.extract_words()
                h = page.height
                # Filter header (~130px) and footer (~60px from bottom)
                filtered = [w for w in words if w['top'] > 130 and w['top'] < h - 60]
                lines = self._group_lines(filtered, tolerance=6)

                for line in lines:
                    line_text = ' '.join(w['text'] for w in line)

                    # Section start marker (only on page 1)
                    if 'TRANSACCIONES' in line_text.upper() and 'CUENTA' in line_text.upper():
                        in_section = True
                        continue
                    if not in_section:
                        continue
                    # Section end
                    if 'RESUMEN DEL BALANCE' in line_text.upper() or 'NOTAS SOBRE' in line_text.upper():
                        if current_block:
                            tx_blocks.append(current_block)
                            current_block = []
                        in_section = False
                        break

                    # Skip column headers line
                    if 'FECHA' in line_text and 'TIPO' in line_text and 'DESCRIPCIÓN' in line_text:
                        continue

                    # New transaction: line contains a day number in FECHA column
                    fecha_words = [w for w in line if w['x0'] < self.X_TIPO
                                   and re.match(r'^\d{1,2}$', w['text'])]
                    if fecha_words:
                        if current_block:
                            tx_blocks.append(current_block)
                        current_block = list(line)
                    else:
                        current_block.extend(line)

        if current_block:
            tx_blocks.append(current_block)

        transactions = []
        prev_balance: Optional[float] = None
        for block in tx_blocks:
            tx, current_balance = self._parse_block(block, prev_balance)
            if current_balance is not None:
                prev_balance = current_balance
            if tx:
                transactions.append(tx)
        return transactions

    def _group_lines(self, words: list[dict], tolerance: int = 6) -> list[list[dict]]:
        """Group words into lines by proximity of top coordinate."""
        if not words:
            return []
        sorted_words = sorted(words, key=lambda w: w['top'])
        lines: list[list[dict]] = []
        current: list[dict] = [sorted_words[0]]
        ref_top = sorted_words[0]['top']

        for w in sorted_words[1:]:
            if abs(w['top'] - ref_top) <= tolerance:
                current.append(w)
            else:
                lines.append(sorted(current, key=lambda x: x['x0']))
                current = [w]
                ref_top = w['top']
        lines.append(sorted(current, key=lambda x: x['x0']))
        return lines

    _MONEY_RE = re.compile(r'^-?€?[\d.]+,\d{2}€?$')

    def _parse_block(self, words: list[dict], prev_balance: Optional[float] = None) -> tuple[Optional[Transaction], Optional[float]]:
        """Parse a transaction from its collected words. Returns (transaction, balance_after).

        The ENTRADA/SALIDA/BALANCE columns are left-aligned at x-offsets that drift a
        few px between PDF exports (and the balance's leading '€' occasionally glues to
        the digits, shifting its x0 left). Rather than trust fixed column boundaries,
        every money-shaped token past X_DESC is collected and sorted by position: the
        rightmost one is always the running balance, the one before it (if any) is the
        transaction amount. Direction (income/expense) is then derived from the row's
        tipo/description text instead of which column the amount happened to render in.
        """
        fecha_words, tipo_words, desc_words, money = [], [], [], []

        for w in words:
            x = w['x0']
            t = w['text']
            if x < self.X_DESC:
                # FECHA/TIPO boundary drifts a few px between PDF exports;
                # classify by content (day/month/year token) instead of x0.
                if re.match(r'^\d{1,2}$', t) or re.match(r'^\d{4}$', t) or t.lower() in MONTH_ES:
                    fecha_words.append(t)
                else:
                    tipo_words.append(t)
            elif self._MONEY_RE.match(t):
                money.append((x, t))
            elif t != '€':
                desc_words.append(t)

        date = _parse_date_es(' '.join(fecha_words))
        if not date:
            return None, None

        tipo = ' '.join(tipo_words).strip()
        desc = _clean_tr_desc(' '.join(desc_words))
        money.sort(key=lambda m: m[0])
        current_balance = _parse_amount(money[-1][1]) if money else None
        amount = _parse_amount(money[-2][1]) if len(money) >= 2 else None

        # Interest / bonuses / dividends → investment return, not received income
        if any(k in tipo for k in ('Interés', 'Bonificación', 'Rentabilidad')):
            if amount:
                return Transaction(date=date, description=desc or tipo, amount=amount,
                                   tx_type='investment', bank='Trade Republic'), current_balance
            return None, current_balance

        # Investment operations
        if 'Operar' in tipo:
            return Transaction(date=date, description=desc, amount=amount or 0,
                               tx_type='investment', bank='Trade Republic'), current_balance

        # Transfers
        if 'Transferencia' in tipo:
            if not amount:
                return None, current_balance
            if _is_internal(desc):
                return Transaction(date=date, description=desc, amount=amount,
                                   tx_type='internal', bank='Trade Republic'), current_balance
            if 'Outgoing' in desc:
                return Transaction(date=date, description=desc, amount=amount,
                                   tx_type='expense', bank='Trade Republic'), current_balance
            if 'Incoming' in desc:
                return Transaction(date=date, description=desc, amount=amount,
                                   tx_type='income', bank='Trade Republic'), current_balance
            # No direction keyword found → fall back to balance movement
            if prev_balance is not None and current_balance is not None and current_balance < prev_balance:
                return Transaction(date=date, description=desc, amount=amount,
                                   tx_type='expense', bank='Trade Republic'), current_balance
            return Transaction(date=date, description=desc, amount=amount,
                               tx_type='income', bank='Trade Republic'), current_balance

        # Card transactions
        if 'tarjeta' in tipo or 'Transacción' in tipo:
            if amount:
                # Use balance direction to distinguish expense from refund.
                if prev_balance is not None and current_balance is not None:
                    if current_balance < prev_balance:
                        return Transaction(date=date, description=desc, amount=amount,
                                           tx_type='expense', bank='Trade Republic'), current_balance
                    else:
                        return Transaction(date=date, description=f'[Devolución] {desc}',
                                           amount=amount, tx_type='income', bank='Trade Republic'), current_balance
                # Fallback: no balance data → assume expense (card transactions are usually expenses)
                return Transaction(date=date, description=desc, amount=amount,
                                   tx_type='expense', bank='Trade Republic'), current_balance
            return None, current_balance

        # Unrecognized tipo: if there's an amount, treat as expense rather than drop silently
        if amount:
            return Transaction(date=date, description=desc or tipo, amount=amount,
                               tx_type='expense', bank='Trade Republic'), current_balance
        return None, current_balance


# ── Openbank XLS (HTML) Parser ─────────────────────────────────────────────────

class OpenbankParser:

    def parse(self, xls_path: str) -> list[Transaction]:
        with open(xls_path, encoding='iso-8859-1') as f:
            content = f.read()

        soup = BeautifulSoup(content, 'html.parser')
        transactions = []
        for row in soup.find_all('tr'):
            tx = self._parse_row(row)
            if tx:
                transactions.append(tx)
        return transactions

    def _parse_row(self, row) -> Optional[Transaction]:
        cells = [td.get_text(strip=True) for td in row.find_all('td')]
        if len(cells) < 8:
            return None

        # Openbank HTML has empty cols between data cols: [''、date、''、date_valor、''、concepto、''、importe、...]
        try:
            date = datetime.strptime(cells[1], '%d/%m/%Y')
        except ValueError:
            return None

        concepto = cells[5] if len(cells) > 5 else ''
        importe_str = cells[7] if len(cells) > 7 else ''

        if not concepto or not importe_str:
            return None

        importe_clean = importe_str.replace('.', '').replace(',', '.').strip()
        try:
            amount = float(importe_clean)
        except ValueError:
            return None

        return _classify_openbank(date, concepto, amount)


# ── Openbank PDF Parser ────────────────────────────────────────────────────────

class OpenbankPDFParser:
    """
    Parse Openbank account statement PDFs.
    Column layout (x0 boundaries):
      Fecha Operación: ~57 | Fecha Valor: ~122 | Concepto: 180–441
      Importe: 441–530 | Saldo: ≥530
    """
    X_FECHA2 = 115
    X_CONCEPTO = 175
    X_IMPORTE = 441
    X_SALDO = 528

    _AMOUNT_RE = re.compile(r'^-?\d{1,3}(?:\.\d{3})*,\d{2}$')
    _DATE_RE = re.compile(r'^\d{2}/\d{2}/\d{4}$')
    _NOISE = re.compile(r'^(EUR|Fecha|Operación|Valor|Concepto|Importe|Saldo|FIN)$', re.I)

    def parse(self, pdf_path: str) -> list[Transaction]:
        transactions = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                words = page.extract_words()
                transactions.extend(self._parse_page(words))
        return transactions

    def _parse_page(self, words: list[dict]) -> list[Transaction]:
        # Find top-y of each transaction's fecha column (x0 ~57, DD/MM/YYYY)
        # Concepto can start ~6px BEFORE the date line, so we use date_top - 6 as block start
        date_tops = sorted({
            w['top'] for w in words
            if self._DATE_RE.match(w['text']) and 50 < w['x0'] < 90
        })
        if not date_tops:
            return []

        # Build blocks: words with top in [date_top_i - 6, date_top_{i+1} - 6)
        results = []
        for i, dt in enumerate(date_tops):
            block_start = dt - 6
            block_end = date_tops[i + 1] - 6 if i + 1 < len(date_tops) else float('inf')
            block = [w for w in words if block_start <= w['top'] < block_end]
            tx = self._parse_block(block)
            if tx:
                results.append(tx)
        return results

    def _parse_block(self, block: list[dict]) -> Optional[Transaction]:
        all_words = block

        date_words = [w for w in all_words if self._DATE_RE.match(w['text']) and w['x0'] < self.X_FECHA2]
        if not date_words:
            return None
        try:
            date = datetime.strptime(date_words[0]['text'], '%d/%m/%Y')
        except ValueError:
            return None

        importe_words = [w for w in all_words if self.X_IMPORTE <= w['x0'] < self.X_SALDO
                         and self._AMOUNT_RE.match(w['text'])]
        if not importe_words:
            return None
        try:
            amount = float(importe_words[0]['text'].replace('.', '').replace(',', '.'))
        except ValueError:
            return None

        # Cap concepto to the vertical range of the first importe to avoid spill from next tx
        importe_top = importe_words[0]['top']
        concepto_words = [w['text'] for w in all_words
                          if self.X_CONCEPTO <= w['x0'] < self.X_IMPORTE
                          and w['top'] <= importe_top + 6
                          and not self._NOISE.match(w['text'])
                          and not self._DATE_RE.match(w['text'])]
        concepto = ' '.join(concepto_words).strip()
        if not concepto:
            return None

        return _classify_openbank(date, concepto, amount)


# ── Revolut PDF Parser ─────────────────────────────────────────────────────────

class RevolutPDFParser:
    X_DATE = 46
    X_DESC = 190
    X_SALIENTE = 390
    X_ENTRANTE = 468
    X_SALDO = 525

    # Old format: "€15.00" (€ prefix, dot decimal)
    # New format: "10,00€" (€ suffix, comma decimal, optional thousands dot)
    _AMOUNT_RE = re.compile(r'^€[\d.]+$|^[\d.,]+€$')
    _DAY_RE = re.compile(r'^\d{1,2}$')

    @staticmethod
    def _parse_amount_str(s: str) -> float:
        s = s.strip('€')
        if ',' in s:
            s = s.replace('.', '').replace(',', '.')
        return float(s)

    def parse(self, pdf_path: str) -> list[Transaction]:
        transactions = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                transactions.extend(self._parse_page(page.extract_words()))
        return transactions

    def _parse_page(self, words: list[dict]) -> list[Transaction]:
        # Use exact top values (no rounding) to avoid float comparison issues
        day_tops = sorted({
            w['top'] for w in words
            if self._DAY_RE.match(w['text']) and w['x0'] < self.X_DATE
        })
        if not day_tops:
            return []
        results = []
        for i, dt in enumerate(day_tops):
            block_end = day_tops[i + 1] if i + 1 < len(day_tops) else float('inf')
            block = [w for w in words if dt <= w['top'] < block_end]
            tx = self._parse_block(block)
            if tx:
                results.append(tx)
        return results

    def _parse_block(self, block: list[dict]) -> Optional[Transaction]:
        first_top = block[0]['top']
        first_line = [w for w in block if abs(w['top'] - first_top) < 3]

        date_words = sorted([w for w in first_line if w['x0'] < 80], key=lambda x: x['x0'])
        date = _parse_date_es(' '.join(w['text'] for w in date_words[:3]))
        if not date:
            return None

        saliente_words = [w for w in first_line if self.X_SALIENTE <= w['x0'] < self.X_ENTRANTE and self._AMOUNT_RE.match(w['text'])]
        entrante_words = [w for w in first_line if self.X_ENTRANTE <= w['x0'] < self.X_SALDO and self._AMOUNT_RE.match(w['text'])]
        saliente = self._parse_amount_str(saliente_words[0]['text']) if saliente_words else None
        entrante = self._parse_amount_str(entrante_words[0]['text']) if entrante_words else None

        if saliente is None and entrante is None:
            return None

        desc_words = [w['text'] for w in first_line if self.X_DESC <= w['x0'] < self.X_SALIENTE]
        desc = ' '.join(desc_words).strip()
        if not desc:
            return None

        desc_upper = desc.upper()

        if 'RECARGA DE' in desc_upper or _is_internal(desc):
            return Transaction(date=date, description=desc, amount=saliente or entrante or 0,
                               tx_type='internal', bank='Revolut')

        if 'REVOLUT DIGITAL' in desc_upper or 'TRANSFER FROM REVOLUT' in desc_upper:
            return Transaction(date=date, description=desc, amount=entrante or saliente or 0,
                               tx_type='investment', bank='Revolut')

        if saliente:
            return Transaction(date=date, description=desc, amount=saliente,
                               tx_type='expense', bank='Revolut')

        if entrante:
            return Transaction(date=date, description=desc, amount=entrante,
                               tx_type='income', bank='Revolut')

        return None


# ── Openbank Cuentas PDF Parser ────────────────────────────────────────────────

class OpenbankCuentasPDFParser:
    """
    Parse Openbank 'Cuentas - Movimientos' PDFs (exported from web portal).

    The PDF has a rotated sidebar at x < 50 that pollutes extract_text().
    Solution: crop each page to x >= 50 before text extraction.

    Text layout per transaction:
      Pre-desc lines (0–2) → Data line: "YYYY-MM-DD YYYY-MM-DD [inline_desc] amount saldo"
      Post-desc line (0–1) — artifacts like "TARJETA : ...6601 EL YYYY-MM-DD"
    """

    _DATA_LINE_RE = re.compile(
        r'^(\d{4}-\d{2}-\d{2})\s+\d{4}-\d{2}-\d{2}\s*(.*?)\s*(-?\d+(?:\.\d+)?)\s+(-?\d+(?:\.\d+)?)\s*$'
    )
    _NOISE_RE = re.compile(
        r'^(Cuentas\s*-|Fecha\s+descarga|Número\s+de\s+Cuenta|Descripci[oó]n:|Titular:|Saldo:|Lista\s+de\s+Movimientos|Fecha$|Operaci[oó]n\s+Fecha|P[aá]gina:)',
        re.IGNORECASE,
    )
    # Post-desc artifact patterns to skip (card detail lines)
    _POSTDESC_SKIP_RE = re.compile(r'^(TARJETA\s*:|:\s*\.\.\.\d|EL\s+\d{4}-)', re.IGNORECASE)

    def parse(self, pdf_path: str) -> list[Transaction]:
        raw_lines: list[str] = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                cropped = page.crop((50, 0, page.width, page.height))
                text = cropped.extract_text() or ''
                for line in text.splitlines():
                    line = line.strip()
                    if not line:
                        continue
                    if self._NOISE_RE.match(line):
                        continue
                    raw_lines.append(line)

        transactions: list[Transaction] = []
        pending_pre: list[str] = []

        i = 0
        while i < len(raw_lines):
            line = raw_lines[i]
            m = self._DATA_LINE_RE.match(line)
            if m:
                date_str, inline_desc, amount_str, _ = m.groups()
                try:
                    date = datetime.strptime(date_str, '%Y-%m-%d')
                    amount = float(amount_str)
                except ValueError:
                    pending_pre = []
                    i += 1
                    continue

                # Collect post-desc: next line if it's not a data line and not noise.
                # Only when inline_desc is empty — a data line that already carries its
                # own description never legitimately has a continuation line; grabbing
                # one anyway steals the next transaction's pre-desc line instead.
                post_parts: list[str] = []
                if not inline_desc.strip() and i + 1 < len(raw_lines):
                    next_line = raw_lines[i + 1]
                    if not self._DATA_LINE_RE.match(next_line) and not self._NOISE_RE.match(next_line):
                        if not self._POSTDESC_SKIP_RE.match(next_line):
                            post_parts.append(next_line)
                        i += 1  # consume it regardless (avoid it becoming pre-desc of next tx)

                parts = [p for p in pending_pre + ([inline_desc.strip()] if inline_desc.strip() else []) + post_parts if p.strip()]
                description = ' '.join(parts).strip() or 'Sin descripción'
                # Remove card artifact tails
                description = re.sub(r'\s+CON LA TARJETA\s*:?\s*$', '', description, flags=re.I).strip()
                description = re.sub(r',\s*CON LA TARJETA\s*$', '', description, flags=re.I).strip()
                # Remove trailing card/date artifacts: "... : ...6601 EL YYYY-MM-DD"
                description = re.sub(r'\s*:\s*\.\.\.\d{4}\s+EL\s+\d{4}-\d{2}-\d{2}$', '', description).strip()
                description = re.sub(r'\s*\.\.\.\d{4}\s+EL\s+\d{4}-\d{2}-\d{2}$', '', description).strip()

                tx = self._classify(date, description, amount)
                if tx:
                    transactions.append(tx)
                pending_pre = []
            else:
                if not self._NOISE_RE.match(line):
                    pending_pre.append(line)
            i += 1

        return transactions

    def _classify(self, date: datetime, description: str, amount: float) -> Optional[Transaction]:
        return _classify_openbank(date, description, amount)


# ── Openbank Tarjeta PDF Parser ────────────────────────────────────────────────

class OpenbankTarjetaPDFParser:
    """
    Parse Openbank card-movements PDFs ('Tarjetas - Movimientos'), distinct from
    account statements: ISO dates, dot-decimal amounts, descriptions truncated
    to a fixed width by the export.

    Line format: "YYYY-MM-DD HH:MM <concepto> Liquidado -X.XX EUR"
    Sidebar at x < 50 pollutes extract_text() same as OpenbankCuentasPDFParser.
    """

    _LINE_RE = re.compile(
        r'^(\d{4}-\d{2}-\d{2})\s+\d{2}:\d{2}\s+(.+?)\s+Liquidado\s+(-?\d+(?:\.\d+)?)\s*EUR$'
    )

    def parse(self, pdf_path: str) -> list[Transaction]:
        transactions = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                cropped = page.crop((50, 0, page.width, page.height))
                text = cropped.extract_text() or ''
                for line in text.splitlines():
                    m = self._LINE_RE.match(line.strip())
                    if not m:
                        continue
                    date_str, concepto, amount_str = m.groups()
                    try:
                        date = datetime.strptime(date_str, '%Y-%m-%d')
                        amount = float(amount_str)
                    except ValueError:
                        continue
                    tx = _classify_openbank(date, concepto.strip(), amount)
                    if tx:
                        transactions.append(tx)
        return transactions


# ── Santander PDF Parser ───────────────────────────────────────────────────────

class SantanderPDFParser:
    """
    Parse Santander account statement PDFs.

    Layout (text-based extraction):
      - Transaction line: "DD mmm YYYY  <description>  X.XXX,XX€  X.XXX,XX€"
      - Value date line:  "F. valor: DD mmm YYYY  [description continuation]"
      - Description may wrap onto the value-date line and subsequent lines.

    Detection: IBAN contains "0049" (Santander BIC prefix).
    """

    # Regex for a transaction date line: starts with "DD mmm YYYY" followed by text
    _TX_DATE_RE = re.compile(
        r'^(\d{1,2})\s+(ene|feb|mar|abr|may|jun|jul|ago|sept?|oct|nov|dic)\s+(\d{4})\s+(.+)',
        re.IGNORECASE,
    )
    # Regex to find amounts at end of line: one or two "X.XXX,XX€" tokens.
    # Some Santander exports use the Unicode minus sign (−, U+2212) instead of
    # a plain hyphen, which would otherwise flip outgoing amounts to positive.
    _AMOUNTS_RE = re.compile(r'([-−]?\d{1,3}(?:\.\d{3})*,\d{2})€')
    # Value date line
    _FVALOR_RE = re.compile(r'^F\.\s*valor:', re.IGNORECASE)
    # Header / noise lines to skip
    _NOISE_RE = re.compile(
        r'^(Titular|Cuenta|Saldo\s+disponible|Fecha\s+operaci|Movimientos\s+de\s+tu\s+cuenta|Documento\s+a\s+fecha)',
        re.IGNORECASE,
    )

    def parse(self, pdf_path: str) -> list[Transaction]:
        raw_lines: list[str] = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''
                raw_lines.extend(text.splitlines())

        # Build logical transaction blocks: each block = [tx_line, fvalor_line?, ...continuation]
        blocks: list[list[str]] = []
        current: list[str] = []

        for line in raw_lines:
            line = line.strip()
            if not line:
                continue
            if self._NOISE_RE.match(line):
                continue
            if self._TX_DATE_RE.match(line):
                if current:
                    blocks.append(current)
                current = [line]
            elif current:
                current.append(line)

        if current:
            blocks.append(current)

        transactions = []
        for block in blocks:
            tx = self._parse_block(block)
            if tx:
                transactions.append(tx)
        return transactions

    def _parse_block(self, block: list[str]) -> Optional[Transaction]:
        first = block[0]
        m = self._TX_DATE_RE.match(first)
        if not m:
            return None

        day, mon, year, rest = m.groups()
        month = MONTH_ES.get(mon.lower()[:3])  # "sept" → "sep"
        if not month:
            return None
        try:
            date = datetime(int(year), month, int(day))
        except ValueError:
            return None

        # Extract all amount tokens from the first line
        amounts = self._AMOUNTS_RE.findall(rest)
        if not amounts:
            return None

        # Last two amounts are importe and saldo; strip them to get description
        # Remove from right: importe (and saldo if present)
        desc_part = rest
        for amt_str in reversed(amounts[-2:]):
            # Remove the amount + € suffix from the right of desc_part
            desc_part = desc_part[:desc_part.rfind(amt_str + '€')].rstrip()

        # Collect continuation text from remaining lines (skip F.valor line itself but keep its tail)
        for line in block[1:]:
            if self._FVALOR_RE.match(line):
                # Remove "F. valor: DD mmm YYYY" prefix, keep the rest as continuation
                tail = re.sub(r'^F\.\s*valor:\s*\d{1,2}\s+\w+\s+\d{4}\s*', '', line, flags=re.IGNORECASE).strip()
                if tail:
                    desc_part = (desc_part + ' ' + tail).strip()
            else:
                # Pure continuation line
                desc_part = (desc_part + ' ' + line).strip()

        description = ' '.join(desc_part.split())

        # Parse importe (second-to-last or only amount)
        importe_str = amounts[-2] if len(amounts) >= 2 else amounts[-1]
        raw = importe_str.replace('.', '').replace(',', '.').replace('−', '-')
        try:
            amount_val = float(raw)
        except ValueError:
            return None

        amount = abs(amount_val)

        # Determine tx_type
        # Santander account belongs to María Ruisánchez: outgoing movements are internal
        # transfers (between own accounts or personal payments), never tracked as expenses.
        if amount_val > 0:
            tx_type = 'income'
        else:
            tx_type = 'internal'

        return Transaction(
            date=date,
            description=description,
            amount=amount,
            tx_type=tx_type,
            bank='Santander',
        )


# ── BBVA xlsx Parser ───────────────────────────────────────────────────────────

# Identidades propias reconocidas como contraparte en transferencias BBVA
# (cuentas Conjunta / Pablo Cavaller personal, agosto 2026 — compra de vivienda).
# Cualquier transferencia de los 4 tipos de abajo cuya contraparte matchee una
# de estas es reasignación de dinero propio, no ingreso/gasto real.
_BBVA_OWN_IDENTITIES = [
    'PABLO CAVALLER GRAU', 'PABLO CAVALLER', 'PABLO BBVA',
    'MARIA RUISANCHEZ', 'RUISANCHEZ GONZALEZ-BARROS', 'BBVA MERI', 'MERI',
    'MYINVESTO', 'MYINVESTOR', 'ACTIVAR',
]

_BBVA_TRANSFER_CONCEPTS = {
    'TRANSFERENCIA RECIBIDA', 'TRANSFERENCIA REALIZADA',
    'TRASPASO DESDE CUENTA', 'TRASPASO A CUENTA',
}

# Compra de vivienda (evento puntual, julio 2026): abono del préstamo, seguros,
# cheques bancarios y tasación → Tipo 'patrimonio', fuera del cash flow mensual.
_BBVA_VIVIENDA_CONCEPTS = {
    'ADEUDO DE SEGUROS',
    'ABONO POR DISPOSICION DE PRESTAMO/CREDITO',
    'CARGO POR EMISION DE CHEQUE BANCARIO',
}
_BBVA_VIVIENDA_TASACION_PREFIX = 'ADEUDO TECNICOS EN TASACION'

# Hipoteca (coste recurrente, SÍ cuenta en cash flow) → Tipo 'expense'.
_BBVA_HIPOTECA_CONCEPT = 'CARGO POR INTERESES DE PRESTAMO'


def _norm_bbva(s: str) -> str:
    return _strip_accents(str(s or '')).upper().strip()


def _is_bbva_known_counterparty(*texts: str) -> bool:
    """Contraparte conocida (propia) para transferencias BBVA. Combina la lista
    explícita de identidades del negocio con la lista global `OWN_ACCOUNT_KEYWORDS`
    (Trade Republic, Sabadell...) ya usada por el resto de parsers, más un match
    laxo para María cuando el nombre viene truncado de forma distinta a las
    variantes conocidas (p.ej. "Maria Rosalia Ruisanchez Gonz...", sin que
    "Ruisanchez Gonzalez-Barros" ni "Maria Ruisanchez" aparezcan literalmente)."""
    combined = ' '.join(texts)
    n = _norm_bbva(combined)
    if any(_norm_bbva(kw) in n for kw in _BBVA_OWN_IDENTITIES):
        return True
    if 'RUISANCHEZ' in n and 'MARIA' in n:
        return True
    return _is_internal(combined)


class BBVAParser:
    """
    Parse BBVA 'Últimos movimientos' xlsx exports (personal o conjunta, mismo
    formato para ambas — el titular se decide fuera del parser, al llamar a
    `SheetsClient.write_transactions(titular=...)`, igual que con el resto de
    parsers del repo).

    Hoja única 'Informe BBVA', cabecera real localizada dinámicamente (suele
    caer en la fila 5, pero no se asume): F.Valor, Fecha, Concepto, Movimiento,
    Importe, Divisa, Disponible, Divisa, Observaciones (columna A vacía).

    Reglas de negocio (compra de vivienda + hipoteca, julio 2026 — ver PROJECT.md):
    - 'Cargo por intereses de prestamo' → Categoría 'Hipoteca', Tipo 'expense'.
    - 'Adeudo de seguros' / 'Abono por disposicion de prestamo/credito' /
      'Cargo por emision de cheque bancario' / 'Adeudo tecnicos en tasacion...'
      → Categoría 'Compra vivienda', Tipo 'patrimonio' (fuera del cash flow).
    - Cualquiera de los 4 conceptos de transferencia/traspaso cuya contraparte
      sea una identidad propia conocida → Tipo 'internal'.
    - Todo lo demás conserva su signo natural: expense/income.
    """

    _HEADER_LABELS = {'F.Valor', 'Fecha', 'Concepto', 'Movimiento', 'Importe', 'Observaciones'}

    def parse(self, xlsx_path: str) -> list[Transaction]:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb['Informe BBVA'] if 'Informe BBVA' in wb.sheetnames else wb[wb.sheetnames[0]]

        header_row, col_idx = self._find_header(ws)
        if header_row is None:
            return []

        transactions = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
            def get(col_name: str):
                idx = col_idx.get(col_name)
                return row[idx - 1].value if idx else None

            fecha_raw = get('Fecha')
            concepto = str(get('Concepto') or '').strip()
            if not fecha_raw or not concepto:
                continue

            if isinstance(fecha_raw, datetime):
                date = fecha_raw
            else:
                try:
                    date = datetime.strptime(str(fecha_raw).strip(), '%d/%m/%Y')
                except ValueError:
                    continue

            try:
                importe = float(get('Importe'))
            except (TypeError, ValueError):
                continue

            movimiento = str(get('Movimiento') or '').strip()
            observaciones = str(get('Observaciones') or '').strip()

            tx = self._classify(date, concepto, movimiento, importe, observaciones)
            if tx:
                transactions.append(tx)
        return transactions

    def _find_header(self, ws) -> tuple[Optional[int], dict[str, int]]:
        max_scan_row = min(ws.max_row, 15)
        for r in range(1, max_scan_row + 1):
            found: dict[str, int] = {}
            for c in range(1, ws.max_column + 1):
                v = ws.cell(row=r, column=c).value
                if isinstance(v, str) and v.strip() in self._HEADER_LABELS:
                    found[v.strip()] = c
            if {'Fecha', 'Concepto', 'Importe'}.issubset(found.keys()):
                return r, found
        return None, {}

    def _classify(self, date: datetime, concepto: str, movimiento: str,
                  importe: float, observaciones: str) -> Optional[Transaction]:
        concepto_n = _norm_bbva(concepto)
        amount = abs(importe)

        if concepto_n == _BBVA_HIPOTECA_CONCEPT:
            return Transaction(date=date, description=concepto, amount=amount,
                               tx_type='expense', bank='BBVA', category='Hipoteca')

        if concepto_n in _BBVA_VIVIENDA_CONCEPTS or concepto_n.startswith(_BBVA_VIVIENDA_TASACION_PREFIX):
            return Transaction(date=date, description=concepto, amount=amount,
                               tx_type='patrimonio', bank='BBVA', category='Compra vivienda')

        contraparte = observaciones or movimiento
        desc = f"{concepto} - {contraparte}" if contraparte else concepto

        if concepto_n in _BBVA_TRANSFER_CONCEPTS and _is_bbva_known_counterparty(movimiento, observaciones):
            return Transaction(date=date, description=desc, amount=amount,
                               tx_type='internal', bank='BBVA')

        if importe < 0:
            return Transaction(date=date, description=desc, amount=amount,
                               tx_type='expense', bank='BBVA')
        return Transaction(date=date, description=desc, amount=amount,
                           tx_type='income', bank='BBVA')
