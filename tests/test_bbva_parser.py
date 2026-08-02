"""Tests del parser BBVA (xlsx) y de las reglas de negocio de la compra de
vivienda de julio 2026 (ver PROJECT.md, sección "Reglas operativas").

No dependen de los xlsx reales de Downloads/ (no versionados): se construyen
workbooks sintéticos con openpyxl replicando el formato real observado —
hoja única "Informe BBVA", filas 1-4 vacías, cabecera en la fila 5, columna
A vacía — para poder testear la lógica de clasificación de forma aislada.
"""

from __future__ import annotations

import openpyxl
import pytest

from parsers import BBVAParser, disambiguate_duplicates, detect_bank

HEADERS = ['F.Valor', 'Fecha', 'Concepto', 'Movimiento', 'Importe', 'Divisa', 'Disponible', 'Divisa', 'Observaciones']


def _build_bbva_xlsx(tmp_path, rows: list[dict]) -> str:
    """rows: lista de dicts con claves Fecha/Concepto/Movimiento/Importe/Observaciones
    (F.Valor se copia de Fecha; Disponible y Divisa se rellenan con valores dummy)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Informe BBVA'
    ws.cell(row=2, column=4, value='Últimos movimientos')
    ws.cell(row=3, column=4, value='Fecha de generación del informe: 01/08/2026')
    for c, label in enumerate(HEADERS, start=2):
        ws.cell(row=5, column=c, value=label)
    for i, row in enumerate(rows, start=6):
        ws.cell(row=i, column=2, value=row['Fecha'])
        ws.cell(row=i, column=3, value=row['Fecha'])
        ws.cell(row=i, column=4, value=row['Concepto'])
        ws.cell(row=i, column=5, value=row.get('Movimiento', ''))
        ws.cell(row=i, column=6, value=row['Importe'])
        ws.cell(row=i, column=7, value='EUR')
        ws.cell(row=i, column=8, value=row.get('Disponible', 0))
        ws.cell(row=i, column=9, value='EUR')
        ws.cell(row=i, column=10, value=row.get('Observaciones', ''))
    path = str(tmp_path / 'bbva.xlsx')
    wb.save(path)
    return path


def test_detect_bank_recognizes_bbva_filename_pattern():
    assert detect_bank("2026Y-08M-01D-13_04_37-Últimos movimientos.xlsx") == 'bbva'
    assert detect_bank("2026Y-08M-01D-13_04_54-Últimos movimientos.xlsx") == 'bbva'


def test_hipoteca_interest_is_expense_not_patrimonio(tmp_path):
    path = _build_bbva_xlsx(tmp_path, [
        {'Fecha': '31/07/2026', 'Concepto': 'Cargo por intereses de prestamo',
         'Movimiento': 'Comisiones,gastos e inter. pagados', 'Importe': -28.55,
         'Observaciones': '0182-2819-11-0830283617'},
    ])
    txs = BBVAParser().parse(path)
    assert len(txs) == 1
    tx = txs[0]
    assert tx.tx_type == 'expense'
    assert tx.category == 'Hipoteca'
    assert tx.amount == pytest.approx(28.55)


@pytest.mark.parametrize('concepto,importe', [
    ('Adeudo de seguros', -4921.66),
    ('Adeudo de seguros', -4396.30),
    ('Abono por disposicion de prestamo/credito', 425317.96),
    ('Cargo por emision de cheque bancario', -234000),
    ('Adeudo tecnicos en tasacion, s.a.', -273.46),
])
def test_vivienda_concepts_are_patrimonio(tmp_path, concepto, importe):
    path = _build_bbva_xlsx(tmp_path, [
        {'Fecha': '30/07/2026', 'Concepto': concepto, 'Movimiento': 'x', 'Importe': importe},
    ])
    tx = BBVAParser().parse(path)[0]
    assert tx.tx_type == 'patrimonio'
    assert tx.category == 'Compra vivienda'
    assert tx.amount == pytest.approx(abs(importe))


def test_two_identical_cheques_same_day_both_parsed_and_disambiguated(tmp_path):
    path = _build_bbva_xlsx(tmp_path, [
        {'Fecha': '30/07/2026', 'Concepto': 'Cargo por emision de cheque bancario',
         'Movimiento': 'Cheques/pagares/efectos - cargos', 'Importe': -234000, 'Disponible': -387433.46},
        {'Fecha': '30/07/2026', 'Concepto': 'Cargo por emision de cheque bancario',
         'Movimiento': 'Cheques/pagares/efectos - cargos', 'Importe': -234000, 'Disponible': -153433.46},
    ])
    txs = BBVAParser().parse(path)
    assert len(txs) == 2
    assert all(tx.tx_type == 'patrimonio' and tx.amount == 234000 for tx in txs)
    # Antes de disambiguate_duplicates, ambas descripciones son idénticas
    # (colisionarían en el dedupe por fecha+descripción+importe+banco de sheets.py).
    assert txs[0].description == txs[1].description
    disambiguate_duplicates(txs)
    assert txs[0].description != txs[1].description
    assert txs[0].description.endswith('(1/2)')
    assert txs[1].description.endswith('(2/2)')


@pytest.mark.parametrize('concepto,movimiento,observaciones', [
    ('Transferencia recibida', 'Bbva meri', 'Bbva Meri'),
    ('Traspaso desde cuenta', 'Pablo cavaller grau', 'PABLO CAVALLER GRAU'),
    ('Transferencia recibida', 'Transferencia desde myinvesto', 'Transferencia desde MyInvesto'),
    ('Transferencia recibida', 'Ruisanchez gonzalez-barros ma', 'RUISANCHEZ GONZALEZ-BARROS MA'),
    ('Traspaso desde cuenta', 'Activar', 'ACTIVAR'),
    ('Transferencia realizada', 'Pablo cavaller', 'Pablo Cavaller'),
    # Nombre completo truncado de forma distinta a las variantes conocidas
    # ("María Rosalia Ruisánchez Gonz..."), pero inequívocamente ella.
    ('Transferencia recibida', 'María rosalia ruisánchez gonz', 'María Rosalia Ruisánchez Gonz'),
    # Contraparte "Trade Es Pablo" -> Trade Republic, vía _is_internal (OWN_ACCOUNT_KEYWORDS)
    ('Transferencia realizada', 'Trade es pablo', 'Trade Es Pablo'),
])
def test_known_identity_transfers_are_internal(tmp_path, concepto, movimiento, observaciones):
    path = _build_bbva_xlsx(tmp_path, [
        {'Fecha': '01/07/2026', 'Concepto': concepto, 'Movimiento': movimiento,
         'Importe': 100.0, 'Observaciones': observaciones},
    ])
    tx = BBVAParser().parse(path)[0]
    assert tx.tx_type == 'internal'


def test_unknown_counterparty_transfer_is_not_internal(tmp_path):
    path = _build_bbva_xlsx(tmp_path, [
        {'Fecha': '01/07/2026', 'Concepto': 'Transferencia recibida', 'Movimiento': 'Marc Fisa',
         'Importe': 500.0, 'Observaciones': 'Marc Fisa'},
    ])
    tx = BBVAParser().parse(path)[0]
    assert tx.tx_type != 'internal'
    assert tx.tx_type == 'income'


def test_real_salary_income_not_misclassified_as_internal(tmp_path):
    path = _build_bbva_xlsx(tmp_path, [
        {'Fecha': '24/07/2026', 'Concepto': 'Transferencia recibida', 'Movimiento': 'Nomina diverinvest julio',
         'Importe': 2126.41, 'Observaciones': 'Nomina Diverinvest Julio'},
    ])
    tx = BBVAParser().parse(path)[0]
    assert tx.tx_type == 'income'
    assert tx.amount == pytest.approx(2126.41)


def test_header_row_located_dynamically_not_hardcoded_to_row_5(tmp_path):
    """El parser no debe asumir que la cabecera cae siempre en la fila 5."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Informe BBVA'
    ws.cell(row=2, column=4, value='Últimos movimientos')
    for c, label in enumerate(HEADERS, start=2):
        ws.cell(row=8, column=c, value=label)  # cabecera desplazada a fila 8
    ws.cell(row=9, column=2, value='01/07/2026')
    ws.cell(row=9, column=3, value='01/07/2026')
    ws.cell(row=9, column=4, value='Recibo variable')
    ws.cell(row=9, column=5, value='Vodafone')
    ws.cell(row=9, column=6, value=-40.0)
    ws.cell(row=9, column=10, value='Vodafone')
    path = str(tmp_path / 'bbva_shifted.xlsx')
    wb.save(path)

    txs = BBVAParser().parse(path)
    assert len(txs) == 1
    assert txs[0].tx_type == 'expense'
    assert txs[0].amount == pytest.approx(40.0)
